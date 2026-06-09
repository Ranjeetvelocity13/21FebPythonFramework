# =========================================================
# Excel Utility - Read/Write Excel Test Data
# =========================================================
import os
import logging
from openpyxl import load_workbook, Workbook

logger = logging.getLogger(__name__)


class ExcelUtils:

    @staticmethod
    def read_excel(file_path, sheet_name="Sheet1"):

        logger.info(
            f"Reading Excel: {file_path}, Sheet: {sheet_name}"
        )

        try:
            wb = load_workbook(
                file_path,
                read_only=True
            )

            ws = wb[sheet_name]

            data = []

            headers = [
                cell.value
                for cell in ws[1]
            ]

            for row in ws.iter_rows(
                min_row=2,
                values_only=True
            ):

                row_data = dict(zip(headers, row))

                data.append(row_data)

            wb.close()

            logger.info(
                f"Excel data read: {len(data)} rows"
            )

            return data

        except Exception as e:
            logger.error(f"Failed to read Excel: {e}")

            raise

    @staticmethod
    def read_cell(file_path, sheet_name, row, column):

        logger.info(
            f"Reading cell ({row}, {column}) from {file_path}"
        )

        wb = load_workbook(
            file_path,
            read_only=True
        )

        ws = wb[sheet_name]

        value = ws.cell(
            row=row,
            column=column
        ).value

        wb.close()

        logger.info(f"Cell value: {value}")

        return value

    @staticmethod
    def get_row_count(file_path, sheet_name="Sheet1"):

        wb = load_workbook(
            file_path,
            read_only=True
        )

        ws = wb[sheet_name]

        count = ws.max_row - 1

        wb.close()

        logger.info(f"Row count: {count}")

        return count

    @staticmethod
    def get_column_count(file_path, sheet_name="Sheet1"):

        wb = load_workbook(
            file_path,
            read_only=True
        )

        ws = wb[sheet_name]

        count = ws.max_column

        wb.close()

        logger.info(f"Column count: {count}")

        return count

    @staticmethod
    def write_cell(
        file_path,
        sheet_name,
        row,
        column,
        value
    ):

        logger.info(
            f"Writing '{value}' to cell ({row}, {column})"
        )

        wb = load_workbook(file_path)

        ws = wb[sheet_name]

        ws.cell(
            row=row,
            column=column,
            value=value
        )

        wb.save(file_path)

        wb.close()

        logger.info("Cell written successfully")

    @staticmethod
    def create_excel(
        file_path,
        sheet_name,
        headers,
        data_rows
    ):

        logger.info(f"Creating Excel: {file_path}")

        wb = Workbook()

        ws = wb.active

        ws.title = sheet_name

        ws.append(headers)

        for row in data_rows:
            ws.append(row)

        wb.save(file_path)

        wb.close()

        logger.info(
            f"Excel created with {len(data_rows)} rows"
        )